from flask import Flask, render_template, request, send_file, redirect, url_for
import pandas as pd
import os
from werkzeug.utils import secure_filename
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
RESULT_FOLDER = "results"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["deg_file"]
        if file.filename == "":
            return redirect(request.url)

        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        zip_path, styled_excel_filename = process_file(filepath)

        return render_template(
            "download.html",
            zip_filename=os.path.basename(zip_path),
            individual_filename=styled_excel_filename
        )

    return render_template("index.html")

def process_file(filepath):
    deg_df = pd.read_excel(filepath)

    expression_col = next((col for col in deg_df.columns if "log2fc" in col.lower()), None)
    pval_col = next((col for col in deg_df.columns if "p_value" in col.lower()), None)

    if not expression_col or not pval_col:
        raise ValueError(f"DEG file is missing Log2FC or p-value column.\nColumns found: {deg_df.columns.tolist()}")

    deg_df = deg_df.rename(columns={
        expression_col: "Expression",
        pval_col: "P_value",
        "GeneSymbol": "Symbol"
    })

    expression_map = dict(zip(deg_df["Symbol"], deg_df["Expression"]))
    pval_map = dict(zip(deg_df["Symbol"], deg_df["P_value"]))

    genelist_df = pd.read_excel("genelistfinalmaster.xlsx", sheet_name=None)

    styled_excel_filename = "final_output.xlsx"
    styled_excel_path = os.path.join(RESULT_FOLDER, styled_excel_filename)
    zip_path = os.path.join(RESULT_FOLDER, "final_output.zip")

    stacked_blocks = []

    with pd.ExcelWriter(styled_excel_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in genelist_df.items():
            df = sheet_df.copy()

            df["Expression"] = df["Symbol"].map(expression_map)
            df["P_value"] = df["Symbol"].map(pval_map)

            df = df.dropna(subset=["Expression", "P_value"])
            df = df.drop_duplicates(subset=["Symbol", "Expression", "P_value"])

            def assign_color(val):
                try:
                    val = float(val)
                    if val > 0: return "Red"
                    elif val < 0: return "Green"
                    elif val == 0: return "Yellow"
                except:
                    return ""

            df["Color"] = df["Expression"].apply(assign_color)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            label_df = pd.DataFrame([[sheet_name]], columns=["Symbol"])
            header_df = pd.DataFrame(columns=["Symbol", "Expression", "P_value", "Color"])
            stacked_block = pd.concat([label_df, header_df, df], ignore_index=True)
            stacked_blocks.append(stacked_block)

        if stacked_blocks:
            stacked_df = pd.concat(stacked_blocks, ignore_index=True)
            stacked_df.to_excel(writer, sheet_name="Stacked_All", index=False)

    # Styling
    wb = load_workbook(styled_excel_path)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="A8FFB0", end_color="A8FFB0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    blue_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        try:
            expr_col = headers.index("Expression") + 1
        except ValueError:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=expr_col)
            try:
                val = float(cell.value)
                if val > 0:
                    cell.fill = red_fill
                elif val < 0:
                    cell.fill = green_fill
                elif val == 0:
                    cell.fill = yellow_fill
            except:
                continue

    # Highlight label rows in Stacked_All
    if "Stacked_All" in wb.sheetnames:
        ws = wb["Stacked_All"]
        for row in range(1, ws.max_row + 1):
            first_cell = ws.cell(row=row, column=1)
            second_cell = ws.cell(row=row, column=2)
            if first_cell.value and not second_cell.value:
                first_cell.fill = blue_fill
                first_cell.font = first_cell.font.copy(bold=True)
                first_cell.alignment = first_cell.alignment.copy(horizontal="center")

    wb.save(styled_excel_path)

    with zipfile.ZipFile(zip_path, "w") as zipf:
        zipf.write(styled_excel_path, arcname=styled_excel_filename)

    return zip_path, styled_excel_filename

@app.route("/download/<filename>")
def download_file(filename):
    return send_file(os.path.join(RESULT_FOLDER, filename), as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
